import json
import logging
import traceback
from pathlib import Path

from PyQt6.QtWidgets import (
    QWidget,
    QVBoxLayout,
    QPushButton,
    QHBoxLayout,
    QFileDialog,
    QMessageBox,
    QTabWidget,
    QAbstractItemDelegate,
    QAbstractItemView,
    QMenuBar,
    QStatusBar,
    QScrollArea,
    QInputDialog,
)
from PyQt6.QtCore import QTimer, Qt
from PyQt6.QtGui import QPainter, QAction
from PyQt6.QtPrintSupport import QPrinter

import os
from src.domain.db import get_connection, startup_verification
from src.domain.masterdata.repo import MasterDataRepo
from src.domain.migration_runner import run_migrations
from src.domain.aggregator import Aggregator
from src.domain.result.service import (
    calculate_result,
    get_result_snapshot,
    get_expense_rows_for_display,
    get_insurance_by_exp_code_for_scenario,
    get_insurance_by_exp_code_from_ui,
    get_labor_rows_from_ui,
)
from src.domain.scenario_input.service import (
    ScenarioInputValidationError,
    post_scenario_input,
    save_canonical_direct,
    get_scenario_input,
    resolve_scenario_id,
)
from .input_panel import InputPanel
from .summary_panel import SummaryPanel
from .labor_detail_table import LaborDetailTable
from .expense_detail_table import ExpenseDetailTable
from .compare.compare_page import ComparePage
from .job_role_table import (
    JobRoleTable,
    setup_qtable_debug_log,
    COL_JOB_CODE,
    COL_JOB_NAME,
    COL_GRADE,
    COL_WORK_DAYS,
    COL_WORK_HOURS,
    COL_OVERTIME_HOURS,
    COL_HOLIDAY_HOURS,
    COL_HEADCOUNT,
)
from .expense_sub_item_table import (
    ExpenseSubItemTable,
    build_sub_items_by_exp,
    EXP_CODES_FROM_LABOR,
    _sub_item_to_dict,
)
from .state import build_canonical_input, compute_action_state
from .export_helpers import build_job_breakdown_rows, build_top_job_summary, build_detail_job_rows, TOP_N_JOB_ROLES
from .donut_chart import DonutChartWidget
from .settings_dialog import SettingsDialog
from .base_year_panel import BaseYearPanel
from .holiday_work_days_panel import HolidayWorkDaysPanel
from .scenario_manager_dialog import ScenarioManagerDialog

# 4-step workflow: controllers + context (optional integration)
try:
    from app.controllers.context import ScenarioContext
    from app.controllers.aggregate_controller import AggregateController
    from app.controllers.save_controller import SaveController
    from app.services.aggregate_service import AggregateService
    from app.repositories.scenario_repository import ScenarioRepository
    from app.domain.models import ResultSnapshot
    _APP_CONTROLLERS_AVAILABLE = True
except ImportError:
    _APP_CONTROLLERS_AVAILABLE = False


def safe_run_save(fn):
    """저장 동작을 감싸서 앱이 죽지 않게 하고, 예외를 로그+팝업으로 남김."""
    try:
        return fn()
    except Exception as exc:
        msg = f"[SAVE ERROR] {exc}\n\n{traceback.format_exc()}"
        logging.error(msg)
        QMessageBox.critical(None, "저장 실패", msg)
        return None


def commit_table_edit(table) -> None:
    if table is None:
        return
    editor = table.focusWidget()
    if editor is not None and editor is not table:
        table.closeEditor(editor, QAbstractItemDelegate.EndEditHint.SubmitModelCache)
    table.clearFocus()


def extract_table_rows(table, columns: dict) -> list[dict]:
    rows = []
    if table is None:
        return rows
    for row in range(table.rowCount()):
        row_data = {}
        for key, col in columns.items():
            if col >= table.columnCount():
                row_data[key] = ""
                continue
            item = table.item(row, col)
            row_data[key] = item.text() if item is not None else ""
        rows.append(row_data)
    return rows


def save_json(path: str, data) -> None:
    with open(path, "w", encoding="utf-8") as file:
        json.dump(data, file, ensure_ascii=False, indent=2)


def _load_job_inputs_from_json(scenario_dir: Path, scenario_id: str) -> dict | None:
    """scenarios/{scenario_id}_job_roles.json 에서 직무별 인원 입력을 읽어 set_job_inputs 형식으로 반환.
    새 형식(work_days, headcount 등)과 구 형식(remark, use_yn 등) 모두 지원. 파일 없거나 오류 시 None."""
    path = scenario_dir / f"{scenario_id}_job_roles.json"
    if not path.exists():
        return None
    try:
        with open(path, encoding="utf-8") as f:
            rows = json.load(f)
    except (OSError, json.JSONDecodeError):
        return None
    if not isinstance(rows, list) or len(rows) == 0:
        return None
    labor_inputs = {}
    first = rows[0] if rows else {}
    # 새 형식: work_days, work_hours, overtime_hours, holiday_days(휴일근로일수), headcount (숫자)
    if "work_days" in first:
        for row in rows:
            if not isinstance(row, dict):
                continue
            job_code = (row.get("job_code") or "").strip()
            if not job_code:
                continue
            try:
                holiday_days = float(row.get("holiday_days", row.get("holiday_hours", 0)) or 0)
                labor_inputs[job_code] = {
                    "work_days": float(row.get("work_days", 0) or 0),
                    "work_hours": float(row.get("work_hours", 0) or 0),
                    "overtime_hours": float(row.get("overtime_hours", 0) or 0),
                    "holiday_work_days": holiday_days,
                    "headcount": int(float(row.get("headcount", 0) or 0)),
                }
            except (TypeError, ValueError):
                labor_inputs[job_code] = {
                    "work_days": 0.0, "work_hours": 0.0, "overtime_hours": 0.0,
                    "holiday_work_days": 0.0, "headcount": 0,
                }
        return labor_inputs if labor_inputs else None
    # 구 형식: remark=work_days, use_yn=work_hours, sort_order=overtime, headcount=직종문자(무시)
    for row in rows:
        if not isinstance(row, dict):
            continue
        job_code = (row.get("job_code") or "").strip()
        if not job_code:
            continue
        try:
            work_days = float(row.get("remark", 0) or 0)
            work_hours = float(row.get("use_yn", 0) or 0)
            overtime = float(row.get("sort_order", 0) or 0)
        except (TypeError, ValueError):
            work_days, work_hours, overtime = 0.0, 0.0, 0.0
        labor_inputs[job_code] = {
            "work_days": work_days,
            "work_hours": work_hours,
            "overtime_hours": overtime,
            "holiday_work_days": 0.0,
            "headcount": 0,
        }
    return labor_inputs if labor_inputs else None


class MainWindow(QWidget):
    """
    자동집하시설 원가산정 프로그램 UI
    """

    def __init__(self):
        super().__init__()
        self.setWindowTitle("자동집하시설 원가산정 프로그램")

        root_layout = QVBoxLayout(self)
        root_layout.setContentsMargins(0, 0, 0, 0)
        root_layout.setSpacing(0)

        setup_qtable_debug_log("logs/qtable_input_debug.log")
        startup_verification(bool(os.environ.get("COSTCALC_DEBUG")))

        # ── 메뉴바 ──
        menu_bar = QMenuBar()
        settings_menu = menu_bar.addMenu("설정")
        settings_action = QAction("설정 열기...", self)
        settings_action.triggered.connect(self._open_settings)
        settings_menu.addAction(settings_action)

        tools_menu = menu_bar.addMenu("도구")
        wage_compare_action = QAction("노임단가 비교...", self)
        wage_compare_action.triggered.connect(self._open_wage_compare)
        tools_menu.addAction(wage_compare_action)

        scenario_manager_action = QAction("시나리오 관리...", self)
        scenario_manager_action.triggered.connect(self._open_scenario_manager)
        tools_menu.addAction(scenario_manager_action)

        root_layout.addWidget(menu_bar)

        # ── 메인 콘텐츠 ──
        content = QWidget()
        layout = QVBoxLayout(content)

        self.input_panel = InputPanel()
        self.base_year_panel = BaseYearPanel()
        self.job_role_table = JobRoleTable()
        self.expense_sub_item_table = ExpenseSubItemTable()
        self.summary_panel = SummaryPanel()
        self.labor_detail = LaborDetailTable()
        self.expense_detail = ExpenseDetailTable()
        self.donut_chart = DonutChartWidget()

        # 버튼 생성
        self.calculate_button = QPushButton("집계 실행")
        self.calculate_button.setProperty("class", "primary")
        self.save_button = QPushButton("시나리오 저장")
        self.save_as_button = QPushButton("다른 이름으로 저장")
        self.load_button = QPushButton("시나리오 불러오기")
        self.export_pdf_button = QPushButton("요약 PDF 내보내기")
        self.export_excel_button = QPushButton("상세 Excel 내보내기")
        self.export_full_excel_button = QPushButton("양식 Excel 내보내기")

        self.calculate_button.clicked.connect(self.calculate)
        self.save_button.clicked.connect(self.save_scenario)
        self.save_as_button.clicked.connect(self.save_scenario_as)
        self.load_button.clicked.connect(self.load_scenario)
        self.export_pdf_button.clicked.connect(self.export_summary_pdf)
        self.export_excel_button.clicked.connect(self.export_details_excel)
        self.export_full_excel_button.clicked.connect(self._export_full_excel)

        self.scenario_dir = Path(__file__).resolve().parents[2] / "scenarios"
        self.scenario_dir.mkdir(parents=True, exist_ok=True)

        # ── 상단 레이아웃: 좌측입력 | 중앙탭 | 우측차트 ──
        top_row = QHBoxLayout()

        # 좌측 패널 (시나리오명)
        left_scroll = QScrollArea()
        left_scroll.setWidgetResizable(True)
        left_scroll.setWidget(self.input_panel)
        left_scroll.setMaximumWidth(260)
        left_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        top_row.addWidget(left_scroll)

        # 중앙 탭 (입력 + 결과 모두 포함)
        self.tabs = QTabWidget()
        self.tabs.addTab(self.base_year_panel, "기준년도")
        self.tabs.addTab(self.job_role_table, "직무별 인원입력")
        self.holiday_work_days_panel = HolidayWorkDaysPanel()
        self.tabs.addTab(self.holiday_work_days_panel, "휴일근무일수 계산")
        self.tabs.addTab(self.expense_sub_item_table, "경비입력")
        self.tabs.addTab(self.summary_panel, "요약")
        self.tabs.addTab(self.labor_detail, "노무비 상세")
        self.tabs.addTab(self.expense_detail, "경비 상세")
        self.tabs.addTab(ComparePage(), "시나리오 비교")
        self.tabs.currentChanged.connect(self._on_tab_changed)
        top_row.addWidget(self.tabs, 1)

        # 우측 도넛 차트
        top_row.addWidget(self.donut_chart)

        layout.addLayout(top_row)

        # ── 하단 버튼바 ──
        button_row = QHBoxLayout()
        button_row.addWidget(self.calculate_button)
        button_row.addWidget(self.save_button)
        button_row.addWidget(self.save_as_button)
        button_row.addWidget(self.load_button)
        button_row.addWidget(self.export_pdf_button)
        button_row.addWidget(self.export_excel_button)
        button_row.addWidget(self.export_full_excel_button)
        layout.addLayout(button_row)

        root_layout.addWidget(content, 1)

        # ── 상태바 ──
        self.status_bar = QStatusBar()
        self.status_bar.showMessage("준비")
        root_layout.addWidget(self.status_bar)

        # ── 초기화 ──
        # 초기 로드 시 dirty 플래그 강제 리셋
        self.job_role_table.dirty = False
        self.expense_sub_item_table.dirty = False
        self._refresh_master_data("default")
        self.last_aggregator = None
        self.last_scenario_name = ""
        self.last_scenario_id = ""
        self.last_labor_rows = []
        self.last_expense_rows = []
        self._canonical_at_aggregation = None  # 집계 실행 시 사용한 입력값 (시나리오 저장 시 이 값으로 저장)
        self._role_name_map = {}
        self._dirty = False
        self._job_role_debounce_timer = QTimer(self)
        self._job_role_debounce_timer.setSingleShot(True)
        self._job_role_debounce_timer.timeout.connect(self._do_job_role_changed)
        self._expense_detail_debounce_timer = QTimer(self)
        self._expense_detail_debounce_timer.setSingleShot(True)
        self._expense_detail_debounce_timer.timeout.connect(self._on_expense_quantity_or_price_changed)
        self._last_labor_count = -1
        self._last_insurance_count = -1
        self._restoring_snapshot = False  # 불러오기 시 저장된 노무비 상세 복원 직후 자동계산으로 덮어쓰기 방지

        self.input_panel.on_change(self._mark_dirty)
        self.base_year_panel.on_change(self._mark_dirty)
        self.holiday_work_days_panel.on_change(self._mark_dirty)
        self.job_role_table.on_change(self._on_job_role_changed)
        self.expense_sub_item_table.on_change(self._mark_dirty)
        # 경비입력 수정 반영 시: 상태 메시지 + 경비상세 탭 재계산
        self.expense_sub_item_table.apply_edit_applied.connect(self._on_expense_edit_applied)
        # 경비입력 탭 '경비코드 저장'만 연결. 시나리오 전체 저장(save_requested)은 여기서 연결하지 않음 → 집계 없이 경비만 저장 가능
        self.expense_sub_item_table.save_current_expense_requested.connect(self._save_current_expense_only)
        self.expense_sub_item_table.set_fetch_default_sub_items(self._fetch_default_expense_sub_items)
        # 경비입력 수량/단가 변경 시 경비상세 자동 재계산 (디바운스)
        self.expense_sub_item_table.quantity_or_price_changed.connect(self._debounce_expense_detail_refresh)

        # 시나리오 목록에서 선택 시 자동 로드
        self.input_panel.scenario_selected.connect(self._load_scenario_by_id)

        # 4-step workflow: bind context and controllers
        if _APP_CONTROLLERS_AVAILABLE:
            self._ctx = ScenarioContext.get()
            self._ctx.subscribe(self._refresh_button_state)
            self._aggregate_controller = AggregateController(self._ctx, AggregateService())
            self._save_controller = SaveController(self._ctx, ScenarioRepository())
        else:
            self._ctx = None
            self._aggregate_controller = None
            self._save_controller = None

        self._refresh_button_state()
        QTimer.singleShot(0, self._force_editable_inputs)
        QTimer.singleShot(200, self._force_editable_inputs)

    # ── 메뉴 핸들러 ──

    def _open_settings(self) -> None:
        dlg = SettingsDialog(self)
        dlg.exec()

    def _open_wage_compare(self) -> None:
        try:
            from src.domain.wage_manager import WageManager
            from .wage_compare_dialog import WageCompareDialog
            wm = WageManager()
            dlg = WageCompareDialog(wm, self)
            dlg.exec()
        except Exception as exc:
            QMessageBox.warning(self, "도구", f"노임단가 비교를 열 수 없습니다:\n{exc}")

    def _open_scenario_manager(self) -> None:
        """시나리오 관리 대화상자 열기"""
        try:
            dlg = ScenarioManagerDialog(self)
            dlg.exec()

            # 삭제된 시나리오가 있으면 처리
            deleted_scenarios = dlg.get_deleted_scenarios()
            if deleted_scenarios:
                # 현재 로드된 시나리오가 삭제되었는지 확인
                current_scenario_id = self.last_scenario_id
                if current_scenario_id in deleted_scenarios:
                    # 현재 시나리오가 삭제됨 - 초기화
                    self.last_scenario_id = ""
                    self.last_scenario_name = ""
                    self.input_panel.scenario_name.setText("시나리오")
                    self._set_dirty(True)
                    self.last_aggregator = None
                    self.summary_panel.update_summary(
                        Aggregator(0, 0, 0, 0, 0, 0), pdf_grand_total=0
                    )
                    self.labor_detail.update_rows([])
                    self.expense_detail.update_rows([], total_headcount=0)
                    self.summary_panel.update()
                    self.labor_detail.update()
                    self.expense_detail.update()
                    self.donut_chart.update_aggregator(None)

                    # default 시나리오의 마스터 데이터 다시 로드
                    self._refresh_master_data("default")

                    self._refresh_button_state()

                    QMessageBox.information(
                        self,
                        "시나리오 관리",
                        f"현재 로드된 시나리오가 삭제되었습니다.\n화면이 초기화되었습니다."
                    )
        except Exception as exc:
            logging.exception("시나리오 관리 대화상자 오류")
            QMessageBox.critical(self, "오류", f"시나리오 관리 중 오류가 발생했습니다:\n{exc}")

    # ── 테이블 편집 ──

    def _on_tab_changed(self, index: int) -> None:
        # 탭 전환 전에 경비입력 테이블의 편집 내용을 메모리에 보존
        commit_table_edit(self.expense_sub_item_table.table)
        self.expense_sub_item_table._save_table_to_current_exp()
        self._force_editable_inputs()
        # 노무비 상세(5), 경비 상세(6) 탭 선택 시 기존 표시 데이터 복원
        if index == 5:
            self.labor_detail.update_rows(self.last_labor_rows)
            self.labor_detail.update()
        elif index == 6:
            # 경비상세 탭 전환 시 경비입력 메모리 기준으로 재계산
            if self.expense_sub_item_table._sub_items_by_exp:
                self._refresh_expense_detail_from_input()
            else:
                total_headcount = sum(
                    v.get("headcount", 0) for v in self.job_role_table.get_job_inputs().values()
                )
                self.expense_detail.update_rows(self.last_expense_rows, total_headcount=total_headcount)
                self.expense_detail.update()

    def _force_editable_inputs(self) -> None:
        commit_table_edit(self.job_role_table.table)
        commit_table_edit(self.expense_sub_item_table.table)
        # 시그널 차단: _force_editable_full이 None 셀에 setItem 호출 시
        # itemChanged → _on_job_role_changed → _do_job_role_changed 연쇄 방지
        self.job_role_table.table.blockSignals(True)
        try:
            self.job_role_table._force_editable()
        finally:
            self.job_role_table.table.blockSignals(False)
        self.expense_sub_item_table._ensure_editable()

    # ── 집계 ──

    def calculate(self):
        # 예약된 노무비 자동계산이 집계 데이터를 덮어쓰지 않도록 타이머 중단
        self._job_role_debounce_timer.stop()

        if self.job_role_table.is_editing() or self.expense_sub_item_table.table.state() == QAbstractItemView.State.EditingState:
            QMessageBox.information(
                self,
                "집계",
                "테이블 셀 편집이 끝난 뒤 다시 시도하세요. (다른 셀을 클릭하거나 Enter로 편집을 마치세요.)",
            )
            return

        # 입력값 검증 (집계 전)
        values = self.input_panel.get_values()
        overhead_text = self.input_panel.overhead_rate.text().strip()
        profit_text = self.input_panel.profit_rate.text().strip()

        # 빈 값 또는 비숫자 입력 검증
        validation_errors = []
        if not overhead_text:
            validation_errors.append("일반관리비율을 입력하세요.")
        else:
            try:
                overhead_val = float(overhead_text)
                if overhead_val < 0 or overhead_val > 100:
                    validation_errors.append("일반관리비율은 0~100 사이의 값이어야 합니다.")
            except ValueError:
                validation_errors.append("일반관리비율에 올바른 숫자를 입력하세요.")

        if not profit_text:
            validation_errors.append("이윤율을 입력하세요.")
        else:
            try:
                profit_val = float(profit_text)
                if profit_val < 0 or profit_val > 100:
                    validation_errors.append("이윤율은 0~100 사이의 값이어야 합니다.")
            except ValueError:
                validation_errors.append("이윤율에 올바른 숫자를 입력하세요.")

        if validation_errors:
            QMessageBox.warning(
                self,
                "입력 오류",
                "다음 항목을 확인하세요:\n\n" + "\n".join(f"• {err}" for err in validation_errors)
            )
            return

        scenario_name = values.get("scenario_name") or "default"
        scenario_id = self._sanitize_filename(scenario_name)
        if not scenario_id:
            QMessageBox.information(self, "집계", "시나리오명을 입력하세요.")
            return

        if _APP_CONTROLLERS_AVAILABLE and self._aggregate_controller is not None:
            self._run_aggregate_via_controller(scenario_id, scenario_name)
            return

        # Legacy path (no app.controllers)
        ok, scenario_id, scenario_name = self._persist_ui_to_db()
        if not ok:
            # _persist_ui_to_db()에서 이미 오류 메시지 표시됨
            return
        self.last_scenario_name = scenario_name
        self.last_scenario_id = scenario_id
        self._set_dirty(False)
        self.job_role_table.dirty = False
        self.expense_sub_item_table.dirty = False

        try:
            self._refresh_master_data(scenario_id)
            # Get overhead_rate and profit_rate from input panel
            input_values = self.input_panel.get_values()
            overhead_rate = input_values.get("overhead_rate", 0.0)
            profit_rate = input_values.get("profit_rate", 0.0)

            conn = get_connection()
            try:
                result = calculate_result(scenario_id, conn, overhead_rate=overhead_rate, profit_rate=profit_rate)
            finally:
                conn.close()
            self._refresh_master_data(scenario_id, skip_expense_reload=True)
            aggregator = result["aggregator"]
            self.summary_panel.update_summary(aggregator, pdf_grand_total=0)
            self.labor_detail.update_rows(result["labor_rows"])
            total_headcount = sum(
                v.get("headcount", 0) for v in self.job_role_table.get_job_inputs().values()
            )
            self.expense_detail.update_rows(result["expense_rows"], total_headcount=total_headcount)
            self.summary_panel.update()
            self.labor_detail.update()
            self.expense_detail.update()
            self.donut_chart.update_aggregator(aggregator)
            self.last_aggregator = aggregator
            self.last_labor_rows = result["labor_rows"]
            self.last_expense_rows = result["expense_rows"]
            conn2 = get_connection()
            try:
                self._canonical_at_aggregation = get_scenario_input(scenario_id, conn2)
            finally:
                conn2.close()
            if _APP_CONTROLLERS_AVAILABLE and self._ctx is not None:
                self._ctx.set_result_snapshot(ResultSnapshot.from_result_dict(result))
            self._refresh_button_state()
            self.status_bar.showMessage("집계 완료 (현재 입력이 반영되었습니다)")
        except Exception as exc:
            logging.exception("집계 실행 중 오류")
            QMessageBox.critical(
                self,
                "집계 오류",
                f"집계 실행 중 오류가 발생했습니다.\n\n{exc}",
            )
            self.status_bar.showMessage("집계 실패")

    def _run_aggregate_via_controller(self, scenario_id: str, scenario_name: str) -> None:
        """Step 3: Run aggregate via AggregateController (persist -> aggregate -> update UI)."""
        def persist() -> bool:
            ok, sid, sname = self._persist_ui_to_db()
            return ok and bool(sid)

        def on_success(snapshot: "ResultSnapshot") -> None:
            # 예약된 노무비 자동계산이 집계 결과를 덮어쓰지 않도록 타이머 중단
            self._job_role_debounce_timer.stop()
            self.last_scenario_name = self._ctx.scenario_name
            self.last_scenario_id = self._ctx.scenario_id
            self._set_dirty(False)
            self.job_role_table.dirty = False
            self.expense_sub_item_table.dirty = False

            # 집계 결과를 먼저 저장 (경비상세 탭 전환 시 사용)
            agg = snapshot.to_dict()["aggregator"]
            self.last_aggregator = Aggregator(
                labor_total=agg["labor_total"],
                fixed_expense_total=agg["fixed_expense_total"],
                variable_expense_total=agg["variable_expense_total"],
                passthrough_expense_total=agg["passthrough_expense_total"],
                overhead_cost=agg["overhead_cost"],
                profit=agg["profit"],
            )
            self.last_labor_rows = snapshot.labor_rows
            self.last_expense_rows = snapshot.expense_rows

            # 마스터 데이터 새로고침 (경비입력 탭 보험 7종만 갱신, 사용자 편집 보존)
            self._refresh_master_data(self._ctx.scenario_id, skip_expense_reload=True)
            conn = get_connection()
            try:
                self._canonical_at_aggregation = get_scenario_input(self._ctx.scenario_id, conn)
            finally:
                conn.close()

            # UI 업데이트: 요약, 노무비 상세, 경비 상세
            self.summary_panel.update_summary(self.last_aggregator, pdf_grand_total=0)
            self.labor_detail.update_rows(snapshot.labor_rows)
            total_headcount = sum(
                v.get("headcount", 0) for v in self.job_role_table.get_job_inputs().values()
            )
            self.expense_detail.update_rows(snapshot.expense_rows, total_headcount=total_headcount)
            self.summary_panel.update()
            self.labor_detail.update()
            self.expense_detail.update()
            self.donut_chart.update_aggregator(self.last_aggregator)
            self._refresh_button_state()
            self.status_bar.showMessage("집계 완료 (현재 입력이 반영되었습니다)")

        def on_error(msg: str) -> None:
            QMessageBox.critical(self, "집계 오류", msg)
            self.status_bar.showMessage("집계 실패")

        self._aggregate_controller.run(
            scenario_id, scenario_name,
            persist_base_data=persist,
            on_success=on_success,
            on_error=on_error,
        )

    # ── 저장/불러오기 ──

    def save_scenario(self):
        return safe_run_save(self._save_scenario_impl)

    def _save_scenario_impl(self):
        """시나리오별 저장: 집계 실행 후 저장 시 해당 시점의 입력·스냅샷으로 저장. 그 외에는 현재 UI 반영."""
        # 예약된 노무비 자동계산이 저장 데이터를 덮어쓰지 않도록 타이머 중단
        self._job_role_debounce_timer.stop()
        scenario_name_raw = (self.input_panel.get_values().get("scenario_name") or "").strip()
        scenario_name = (scenario_name_raw or "default").strip()
        scenario_id = self._sanitize_filename(scenario_name_raw or "default")
        if not scenario_id:
            QMessageBox.information(self, "저장", "시나리오명을 입력하세요.")
            return

        has_snapshot = (
            _APP_CONTROLLERS_AVAILABLE
            and self._save_controller is not None
            and self._ctx is not None
            and self._ctx.result_snapshot is not None
        )
        use_aggregation_state = has_snapshot and self._canonical_at_aggregation is not None

        if use_aggregation_state:
            # 집계 실행 직후 저장: 집계 시 사용한 입력값(canonical)으로 저장해 불러오기 시 동일 값 복원
            conn = get_connection()
            try:
                run_migrations(conn)
                repo = MasterDataRepo(conn)
                repo.ensure_job_roles_for_scenario(scenario_id)
                repo.ensure_expense_masterdata_for_scenario(scenario_id)
                canon = dict(self._canonical_at_aggregation)
                canon["_display_name"] = scenario_name  # 재실행 후 목록/불러오기 시 표시명 복원용
                save_canonical_direct(conn, scenario_id, canon)
                labor_roles = self._canonical_at_aggregation.get("labor", {}).get("job_roles", {})
                job_roles_json = []
                for code, vals in labor_roles.items():
                    if not isinstance(vals, dict):
                        continue
                    job_roles_json.append({
                        "job_code": code,
                        "work_days": vals.get("work_days", 0),
                        "work_hours": vals.get("work_hours", 0),
                        "overtime_hours": vals.get("overtime_hours", 0),
                        "holiday_days": vals.get("holiday_work_days", vals.get("holiday_work_hours", 0)),
                        "headcount": vals.get("headcount", 0),
                    })
                save_json(str(self.scenario_dir / f"{scenario_id}_job_roles.json"), job_roles_json)
                # 경비 세부항목도 DB에 저장 (집계 후에도 사용자 편집값 반영)
                commit_table_edit(self.expense_sub_item_table.table)
                total_headcount = sum(
                    v.get("headcount", 0) for v in self.job_role_table.get_job_inputs().values()
                )
                all_sub = self.expense_sub_item_table.get_all_sub_items(total_headcount=total_headcount)
                logging.info("[저장-집계] 경비 세부항목 %d개 코드 저장 시작", len(all_sub))
                for exp_code, sub_list in all_sub.items():
                    for si in sub_list:
                        logging.info("[저장-집계] %s: sub=%s qty=%s price=%s amount=%s",
                                     exp_code, si.get("sub_code"), si.get("quantity"), si.get("unit_price"), si.get("amount"))
                    repo.upsert_expense_sub_items(scenario_id, exp_code, sub_list)
            except ScenarioInputValidationError as exc:
                self._show_validation_errors(exc.errors)
                return
            except Exception as exc:
                logging.exception("저장 중 오류")
                QMessageBox.critical(self, "저장 실패", str(exc))
                return
            finally:
                conn.close()
        else:
            ok, scenario_id, scenario_name = self._persist_ui_to_db()
            if not ok:
                return

        self.last_scenario_name = scenario_name
        self.last_scenario_id = scenario_id
        self._set_dirty(False)
        self.job_role_table.dirty = False
        self.expense_sub_item_table.dirty = False

        snapshot_saved = False
        if has_snapshot:
            try:
                self._save_controller._repo.save_scenario_with_snapshot(
                    scenario_id, scenario_name, self._ctx.result_snapshot
                )
                snapshot_saved = True
            except Exception as exc:
                logging.exception("저장 중 오류")
                QMessageBox.critical(self, "저장 실패", str(exc))
                return

        # 시나리오 목록 갱신 (시그널 차단하여 불필요한 재계산 방지)
        self.job_role_table.table.blockSignals(True)
        try:
            self.input_panel.refresh_scenario_list()
        finally:
            self.job_role_table.table.blockSignals(False)

        if snapshot_saved:
            QMessageBox.information(self, "저장 완료", "시나리오(직무·경비·집계 결과)가 저장되었습니다.")
        else:
            QMessageBox.information(
                self,
                "저장 완료",
                "시나리오 입력(직무·경비)이 저장되었습니다.\n집계 결과를 반영하려면 '집계 실행' 후 다시 저장하세요.",
            )

    def save_scenario_as(self):
        """다른 이름으로 시나리오 저장"""
        # 현재 시나리오명 가져오기
        current_name = self.input_panel.get_values().get("scenario_name") or ""

        # 새 시나리오명 입력 받기
        new_name, ok = QInputDialog.getText(
            self,
            "다른 이름으로 저장",
            "새 시나리오명을 입력하세요:",
            text=current_name + "_복사본" if current_name else ""
        )

        if not ok or not new_name.strip():
            return

        new_name = new_name.strip()

        # 시나리오명 유효성 검사
        new_scenario_id = self._sanitize_filename(new_name)
        if not new_scenario_id:
            QMessageBox.warning(
                self,
                "입력 오류",
                "유효한 시나리오명을 입력하세요.\n특수문자는 사용할 수 없습니다."
            )
            return

        # 기존 시나리오와 이름이 같은지 확인
        conn = get_connection()
        try:
            from src.domain.scenario_input.service import list_scenario_ids
            existing_scenarios = list_scenario_ids(conn)

            if new_scenario_id in existing_scenarios:
                reply = QMessageBox.question(
                    self,
                    "시나리오 덮어쓰기",
                    f"'{new_name}' 시나리오가 이미 존재합니다.\n덮어쓰시겠습니까?",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                    QMessageBox.StandardButton.No
                )

                if reply != QMessageBox.StandardButton.Yes:
                    return
        finally:
            conn.close()

        # 시나리오명 임시 변경
        original_name = self.input_panel.scenario_name.text()
        self.input_panel.scenario_name.setText(new_name)

        # 저장 실행
        try:
            self.save_scenario()
        except Exception as e:
            # 오류 발생 시 원래 이름으로 복원
            self.input_panel.scenario_name.setText(original_name)
            QMessageBox.critical(
                self,
                "저장 실패",
                f"시나리오 저장 중 오류가 발생했습니다.\n{e}"
            )

    def _load_scenario_by_id(self, scenario_id: str):
        """시나리오 목록에서 선택된 시나리오 로드"""
        self.input_panel.scenario_name.setText(scenario_id)
        self.load_scenario()

    def load_scenario(self):
        # 셀 편집 중이면 편집을 먼저 확정
        commit_table_edit(self.job_role_table.table)
        commit_table_edit(self.expense_sub_item_table.table)
        if self.job_role_table.is_editing() or self.expense_sub_item_table.table.state() == QAbstractItemView.State.EditingState:
            QMessageBox.information(
                self,
                "시나리오 불러오기",
                "테이블 셀 편집이 끝난 뒤 다시 시도하세요. (다른 셀을 클릭하거나 Enter로 편집을 마치세요.)",
            )
            return
        if self.job_role_table.dirty or self.expense_sub_item_table.dirty:
            reply = QMessageBox.question(
                self,
                "시나리오 불러오기",
                "저장되지 않은 변경사항이 있습니다.\n\n변경사항을 무시하고 불러오시겠습니까?\n(예: 변경 무시 후 불러오기 / 아니오: 취소)",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No,
            )
            if reply != QMessageBox.StandardButton.Yes:
                return
            # 사용자가 "예"를 선택하면 dirty 플래그 강제 초기화
            self.job_role_table.dirty = False
            self.expense_sub_item_table.dirty = False
        values = self.input_panel.get_values()
        scenario_name = (values.get("scenario_name") or "").strip() or "default"
        if not scenario_name:
            QMessageBox.information(self, "불러오기", "시나리오명을 입력하세요.")
            return

        # 불러오기 중 예약된 노무비 자동계산이 저장된 스냅샷을 덮어쓰지 않도록 타이머 중단
        self._job_role_debounce_timer.stop()

        # 현재 입력된 직무별 데이터를 임시 저장 (마스터 데이터 새로고침 전)
        current_job_inputs = self.job_role_table.get_job_inputs()

        conn = get_connection()
        try:
            # '2023설계' 등 한글 제거 시 id가 '2023'이 되어도 저장된 '2023_' 시나리오를 찾도록 보정
            scenario_id = resolve_scenario_id(conn, scenario_name, self._sanitize_filename)
            canonical = get_scenario_input(scenario_id, conn)
        finally:
            conn.close()

        self._refresh_master_data(scenario_id)

        # 저장된 시나리오 데이터가 있으면 적용, 없으면 JSON 백업 또는 현재 입력값 유지
        labor_inputs = canonical.get("labor", {}).get("job_roles", {})
        if labor_inputs:
            self._apply_canonical_input(canonical)
        else:
            # DB에 없으면 scenarios/*_job_roles.json 에서 직무별 인원 복원 시도
            fallback = _load_job_inputs_from_json(self.scenario_dir, scenario_id)
            to_apply = fallback if fallback else current_job_inputs
            self.job_role_table.table.blockSignals(True)
            try:
                self.job_role_table.set_job_inputs(to_apply)
            finally:
                self.job_role_table.table.blockSignals(False)

        self.last_scenario_id = scenario_id
        # 저장된 표시명이 있으면 시나리오명 필드를 그 값으로 맞춤 (재실행 후 일관된 표시)
        saved_display = canonical.get("_display_name") or scenario_name
        self.last_scenario_name = saved_display
        if saved_display and saved_display != scenario_name:
            self.input_panel.scenario_name.setText(saved_display)
        self._canonical_at_aggregation = None  # 불러오기 시 초기화 (다음 집계 후 저장 시 새로 설정됨)
        if _APP_CONTROLLERS_AVAILABLE and self._ctx is not None:
            self._ctx.set_scenario(scenario_id, scenario_name)
            self._ctx.set_dirty(False)
        self._set_dirty(False)
        self.job_role_table._force_editable()

        # 불러온 시나리오에 저장된 집계 결과가 있으면 경비 상세·노무비 상세 복원
        conn2 = get_connection()
        try:
            snapshot = get_result_snapshot(scenario_id, conn2)
        finally:
            conn2.close()
        if snapshot:
            agg = snapshot.get("aggregator", {})
            aggregator = Aggregator(
                labor_total=agg.get("labor_total", 0),
                fixed_expense_total=agg.get("fixed_expense_total", 0),
                variable_expense_total=agg.get("variable_expense_total", 0),
                passthrough_expense_total=agg.get("passthrough_expense_total", 0),
                overhead_cost=agg.get("overhead_cost", 0),
                profit=agg.get("profit", 0),
            )
            self.last_aggregator = aggregator
            self.last_labor_rows = snapshot.get("labor_rows", [])
            self.last_expense_rows = snapshot.get("expense_rows", [])
            if _APP_CONTROLLERS_AVAILABLE and self._ctx is not None:
                self._ctx.set_result_snapshot(ResultSnapshot.from_result_dict(snapshot))
            self.summary_panel.update_summary(aggregator, pdf_grand_total=0)
            self.labor_detail.update_rows(self.last_labor_rows)

            # 인원 합계 계산
            total_headcount = sum(
                v.get("headcount", 0) for v in self.job_role_table.get_job_inputs().values()
            )
            self.expense_detail.update_rows(self.last_expense_rows, total_headcount=total_headcount)

            self.summary_panel.update()
            self.labor_detail.update()
            self.expense_detail.update()
            self.donut_chart.update_aggregator(aggregator)
            self._refresh_button_state()
            # 불러오기 직후 _do_job_role_changed가 호출되어 저장된 노무비 상세를 덮어쓰는 것 방지
            self._restoring_snapshot = True
            QTimer.singleShot(400, self._clear_restoring_snapshot)
        else:
            self.last_aggregator = None
            self.last_labor_rows = []
            self.last_expense_rows = []
            if _APP_CONTROLLERS_AVAILABLE and self._ctx is not None:
                self._ctx.clear_after_load()
            self.summary_panel.update_summary(
                Aggregator(0, 0, 0, 0, 0, 0), pdf_grand_total=0
            )
            self.labor_detail.update_rows([])

            # 인원 합계 계산
            total_headcount = sum(
                v.get("headcount", 0) for v in self.job_role_table.get_job_inputs().values()
            )
            self.expense_detail.update_rows([], total_headcount=total_headcount)

            self.summary_panel.update()
            self.labor_detail.update()
            self.expense_detail.update()
            self.donut_chart.update_aggregator(None)
            # 스냅샷 없을 때도 직무별 입력 기준으로 노무비 상세 자동계산 한 번 실행
            QTimer.singleShot(0, self._on_job_role_changed)
            self._refresh_button_state()

    def _sanitize_filename(self, name: str) -> str:
        safe = "".join(ch for ch in name if ch.isalnum() or ch in ("-", "_"))
        return safe or "scenario"

    def _persist_ui_to_db(self):
        """현재 UI 입력을 DB에 반영. 성공 시 (True, scenario_id, scenario_name), 실패 시 (False, None, None)."""
        # 예약된 노무비 자동계산이 저장할 데이터를 덮어쓰지 않도록 타이머 중단
        self._job_role_debounce_timer.stop()
        commit_table_edit(self.job_role_table.table)
        commit_table_edit(self.expense_sub_item_table.table)
        values = self.input_panel.get_values()
        scenario_name = (values.get("scenario_name") or "").strip() or "default"
        scenario_id = self._sanitize_filename(scenario_name)
        if not scenario_id:
            return False, None, None
        logging.info("[저장] 시나리오 저장 대상 id=%r 이름=%r", scenario_id, scenario_name)

        data = extract_table_rows(
            self.job_role_table.table,
            {
                "job_code": COL_JOB_CODE,
                "job_name": COL_JOB_NAME,
                "grade": COL_GRADE,
                "work_days": COL_WORK_DAYS,
                "work_hours": COL_WORK_HOURS,
                "overtime_hours": COL_OVERTIME_HOURS,
                "holiday_days": COL_HOLIDAY_HOURS,
                "headcount": COL_HEADCOUNT,
            },
        )
        logging.info("[저장] 직무행 수 = %s", len(data))
        save_json(str(self.scenario_dir / f"{scenario_id}_job_roles.json"), data)

        conn = get_connection()
        try:
            run_migrations(conn)
            repo = MasterDataRepo(conn)
            # 새 시나리오: 직무/경비 마스터가 없으면 default에서 복사 → 불러오기 시 저장 데이터 있음
            repo.ensure_job_roles_for_scenario(scenario_id)
            repo.ensure_expense_masterdata_for_scenario(scenario_id)
            table_rows = self.job_role_table.table.rowCount()
            job_inputs = self.job_role_table.get_job_inputs()
            logging.info("[저장] 직무 테이블 행=%d, UI 직무입력=%d건", table_rows, len(job_inputs))
            if len(job_inputs) > 0:
                logging.info("[저장] 직무입력 샘플: %s", list(job_inputs.items())[:2])
            # 경비 탭: 저장 대상이 현재 불러온 시나리오면 메모리 전체 저장; 아니면 DB 기준으로 현재 경비코드만 반영
            commit_table_edit(self.expense_sub_item_table.table)
            total_headcount = sum(
                v.get("headcount", 0) for v in job_inputs.values()
            ) or 1
            est = self.expense_sub_item_table
            est._total_headcount = max(int(total_headcount), 1)
            est._save_table_to_current_exp()

            if getattr(self, "last_scenario_id", None) == scenario_id:
                # 방금 불러온 시나리오와 동일 → 수정한 모든 경비코드 반영
                all_sub = est.get_all_sub_items(total_headcount=total_headcount)
            else:
                # 다른 시나리오 또는 한 번도 불러오지 않음 → DB에 있던 값 유지, 메모리에 있는 모든 경비코드 반영
                existing_raw = repo.get_expense_sub_items(scenario_id)
                existing_by_exp = {}
                for si in existing_raw:
                    existing_by_exp.setdefault(si.exp_code, []).append(_sub_item_to_dict(si))
                if existing_by_exp:
                    # 메모리에 있는 모든 경비코드의 수정사항을 DB 데이터 위에 덮어쓰기
                    for mem_code, mem_items in est._sub_items_by_exp.items():
                        if mem_code not in EXP_CODES_FROM_LABOR and mem_items:
                            existing_by_exp[mem_code] = list(mem_items)
                    all_sub = {k: v for k, v in existing_by_exp.items() if k not in EXP_CODES_FROM_LABOR}
                else:
                    all_sub = est.get_all_sub_items(total_headcount=total_headcount)

            expense_inputs = {}
            for exp_code, sub_list in all_sub.items():
                if not sub_list:
                    continue
                total_qty = sum(float(si.get("quantity", 0) or 0) for si in sub_list)
                total_amt = sum(int(si.get("amount", 0) or 0) for si in sub_list)
                unit_price = int(total_amt / total_qty) if total_qty else 0
                expense_inputs[exp_code] = {"quantity": total_qty, "unit_price": unit_price}
            logging.info("[저장] 경비 입력 %d개 코드 canonical 반영 (기준: DB 병합)", len(expense_inputs))
            ui_data = {
                "job_inputs": job_inputs,
                "expense_inputs": expense_inputs,
            }
            input_values = self.input_panel.get_values()
            overhead_rate = input_values.get("overhead_rate", 0.0)
            profit_rate = input_values.get("profit_rate", 0.0)
            year_values = self.base_year_panel.get_values()
            base_year = year_values.get("base_year")
            wage_year = year_values.get("wage_year")
            wage_half = year_values.get("wage_half")
            holiday_calc = self.holiday_work_days_panel.get_values()
            payload = build_canonical_input(
                ui_data["job_inputs"],
                ui_data["expense_inputs"],
                overhead_rate=overhead_rate,
                profit_rate=profit_rate,
                base_year=base_year,
                wage_year=wage_year,
                wage_half=wage_half,
                holiday_work_days_calc=holiday_calc,
            )
            payload["_display_name"] = scenario_name  # 재실행 후 목록/불러오기 시 표시명 복원용
            logging.info("[저장] canonical 직무수: %d건, 경비: %d건",
                        len(payload.get("labor", {}).get("job_roles", {})),
                        len(payload.get("expenses", {}).get("items", {})))
            post_scenario_input(payload, scenario_id, conn)
            logging.info("[저장-직접] 경비 세부항목 %d개 코드 DB 저장 시작", len(all_sub))
            for exp_code, sub_list in all_sub.items():
                for si in sub_list:
                    logging.info("[저장-직접] %s: sub=%s qty=%s price=%s amount=%s",
                                 exp_code, si.get("sub_code"), si.get("quantity"), si.get("unit_price"), si.get("amount"))
                repo.upsert_expense_sub_items(scenario_id, exp_code, sub_list)
            return True, scenario_id, scenario_name
        except ScenarioInputValidationError as exc:
            self._show_validation_errors(exc.errors)
            return False, None, None
        finally:
            conn.close()

    def _save_current_expense_only(self, *, silent: bool = False) -> None:
        """경비입력 탭에서 선택한 경비코드의 세부 항목만 DB에 저장. 집계 실행과 별개.
        silent=True: 수량/단가 자동 저장 시 사용. 조건 미충족·실패 시 메시지 박스 없이 반환."""
        commit_table_edit(self.expense_sub_item_table.table)
        values = self.input_panel.get_values()
        scenario_name = values.get("scenario_name") or "default"
        scenario_id = self._sanitize_filename(scenario_name)
        if not scenario_id:
            if not silent:
                QMessageBox.information(self, "저장", "시나리오명을 입력하세요.")
            return
        total_headcount = sum(
            v.get("headcount", 0) for v in self.job_role_table.get_job_inputs().values()
        )
        exp_code, sub_list = self.expense_sub_item_table.get_current_exp_sub_items(
            total_headcount=max(total_headcount, 1)
        )
        if not exp_code:
            if not silent:
                QMessageBox.information(
                    self, "저장", "저장할 경비코드를 선택한 뒤 '경비코드 선택'을 눌러 표시하고 저장하세요."
                )
            return
        try:
            conn = get_connection()
            try:
                repo = MasterDataRepo(conn)
                repo.ensure_expense_masterdata_for_scenario(scenario_id)
                repo.upsert_expense_sub_items(scenario_id, exp_code, sub_list)
            finally:
                conn.close()
            if not silent:
                self.status_bar.showMessage(f"경비코드 '{exp_code}' 저장됨 (해당 항목만 반영)")
        except Exception as exc:
            logging.exception("경비코드 저장 오류")
            if not silent:
                QMessageBox.critical(self, "저장 실패", f"경비코드 저장 중 오류가 발생했습니다.\n{exc}")

    def _fetch_default_expense_sub_items(self, exp_code: str):
        """경비코드 선택 시 해당 코드의 기본 세부 항목을 default 시나리오에서 불러온다. (exp_code별 세부 항목·단가 등)"""
        conn = get_connection()
        try:
            repo = MasterDataRepo(conn)
            repo.ensure_expense_masterdata_for_scenario("default")
            return repo.get_expense_sub_items("default", exp_code)
        finally:
            conn.close()

    def _refresh_master_data(self, scenario_id: str, skip_expense_reload: bool = False) -> None:
        if self.job_role_table.is_editing() or self.expense_sub_item_table.table.state() == QAbstractItemView.State.EditingState:
            logging.warning("[새로고침] 편집 중/더티 상태라 생략")
            return
        if self.job_role_table.dirty or self.expense_sub_item_table.dirty:
            logging.warning("[새로고침] 편집 중/더티 상태라 생략")
            return
        commit_table_edit(self.job_role_table.table)
        commit_table_edit(self.expense_sub_item_table.table)
        _ = self.job_role_table.get_job_inputs()
        conn = get_connection()
        repo = MasterDataRepo(conn)
        try:
            repo.ensure_expense_masterdata_for_scenario(scenario_id)
            roles = repo.get_job_roles(scenario_id)
            items = repo.get_expense_items(scenario_id)
            self._role_name_map = {r.job_code: r.job_name for r in roles}
            self.job_role_table.table.blockSignals(True)
            try:
                self.job_role_table.load_roles(
                    [{"job_code": r.job_code, "job_name": r.job_name} for r in roles],
                    default_work_days=20.6,
                    default_work_hours=8.0,
                )
                self.job_role_table.set_available_roles(
                    [{"job_code": r.job_code, "job_name": r.job_name} for r in roles]
                )
                # 저장된 직무별 인원 입력 복원 (load_roles가 headcount를 0으로 초기화하므로)
                canonical = get_scenario_input(scenario_id, conn)
                labor_inputs = canonical.get("labor", {}).get("job_roles", {})
                if labor_inputs:
                    self.job_role_table.set_job_inputs(labor_inputs)
                expense_items_for_sub = [
                    {
                        "exp_code": i.exp_code,
                        "exp_name": i.exp_name,
                        "group_code": i.group_code,
                        "sort_order": i.sort_order,
                    }
                    for i in items
                ]
                if skip_expense_reload:
                    # 집계 직후: 사용자 편집 보존, 보험 7종만 갱신
                    snapshot = get_result_snapshot(scenario_id, conn)
                    labor_insurance = (snapshot.get("insurance_by_exp_code") or {}) if snapshot else None
                    if not labor_insurance:
                        labor_insurance = get_insurance_by_exp_code_for_scenario(scenario_id, conn)
                    if labor_insurance:
                        item_map = {
                            i["exp_code"]: i.get("exp_name", i["exp_code"])
                            for i in expense_items_for_sub
                        }
                        from src.ui.expense_sub_item_table import _labor_insurance_row
                        for exp_code in EXP_CODES_FROM_LABOR:
                            amount = labor_insurance.get(exp_code, 0) or 0
                            self.expense_sub_item_table._sub_items_by_exp[exp_code] = [
                                _labor_insurance_row(exp_code, item_map.get(exp_code, exp_code), int(amount))
                            ]
                    logging.info("[새로고침] 집계 후 경비입력 테이블 리로드 생략 (사용자 편집 보존)")
                else:
                    snapshot = get_result_snapshot(scenario_id, conn)
                    labor_insurance = (snapshot.get("insurance_by_exp_code") or {}) if snapshot else None
                    if not labor_insurance:
                        labor_insurance = get_insurance_by_exp_code_for_scenario(scenario_id, conn)
                    sub_items_by_exp = build_sub_items_by_exp(
                        repo, scenario_id,
                        labor_insurance_by_exp_code=labor_insurance or None,
                        expense_items=expense_items_for_sub,
                    )
                    total_headcount = sum(
                        v.get("headcount", 0) for v in self.job_role_table.get_job_inputs().values()
                    )
                    self.expense_sub_item_table.load_sub_items(
                        sub_items_by_exp, expense_items_for_sub, total_headcount=total_headcount
                    )
            finally:
                self.job_role_table.table.blockSignals(False)
        finally:
            conn.close()

    # ── 내보내기 ──

    def export_summary_pdf(self):
        return safe_run_save(self._export_summary_pdf_impl)

    def _export_summary_pdf_impl(self):
        if self._dirty:
            QMessageBox.information(self, "내보내기", "저장되지 않은 변경사항이 있습니다.")
            return
        snapshot = get_result_snapshot(self.last_scenario_id)
        if snapshot is None:
            QMessageBox.information(self, "내보내기", "계산 결과가 없습니다. 먼저 집계를 실행하세요.")
            return

        safe_name = self._sanitize_filename(self.last_scenario_name)
        default_path = self.scenario_dir / f"{safe_name}_summary.pdf"
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "요약 PDF 내보내기",
            str(default_path),
            "PDF Files (*.pdf)",
        )
        if not file_path:
            return

        printer = QPrinter()
        printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
        printer.setOutputFileName(file_path)

        painter = QPainter(printer)
        y = 40
        line_height = 24

        def draw_line(text: str):
            nonlocal y
            painter.drawText(40, y, text)
            y += line_height

        aggregator = snapshot.get("aggregator", {})
        draw_line(f"시나리오: {self.last_scenario_name}")
        draw_line("집계 요약")
        draw_line(f"노무비 합계: {aggregator.get('labor_total', 0):,}")
        draw_line(f"고정경비 합계: {aggregator.get('fixed_expense_total', 0):,}")
        draw_line(f"변동경비 합계: {aggregator.get('variable_expense_total', 0):,}")
        draw_line(f"대행비 합계: {aggregator.get('passthrough_expense_total', 0):,}")
        draw_line(f"일반관리비: {aggregator.get('overhead_cost', 0):,}")
        draw_line(f"이윤: {aggregator.get('profit', 0):,}")
        draw_line(f"총계(VAT 제외): {aggregator.get('grand_total', 0):,}")

        draw_line("")
        draw_line(f"직무별 노무비 상위 {TOP_N_JOB_ROLES}")
        for job_name, total, _ in build_top_job_summary(snapshot, TOP_N_JOB_ROLES):
            draw_line(f"{job_name}: {total:,}")

        painter.end()

        detail_path = file_path.replace("_summary.pdf", "_detail.pdf")
        self._export_detail_pdf(detail_path, snapshot)
        QMessageBox.information(self, "내보내기", "PDF 저장이 완료되었습니다.")

    def export_details_excel(self):
        return safe_run_save(self._export_details_excel_impl)

    def _export_details_excel_impl(self):
        if self._dirty:
            QMessageBox.information(self, "내보내기", "저장되지 않은 변경사항이 있습니다.")
            return
        snapshot = get_result_snapshot(self.last_scenario_id)
        if snapshot is None:
            QMessageBox.information(self, "내보내기", "계산 결과가 없습니다. 먼저 집계를 실행하세요.")
            return

        try:
            import openpyxl
        except ImportError:
            QMessageBox.warning(
                self,
                "내보내기 실패",
                "openpyxl이 설치되어 있지 않습니다. `pip install openpyxl` 실행 후 다시 시도하세요.",
            )
            return

        safe_name = self._sanitize_filename(self.last_scenario_name)
        default_path = self.scenario_dir / f"{safe_name}_details.xlsx"
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "상세 Excel 내보내기",
            str(default_path),
            "Excel Files (*.xlsx)",
        )
        if not file_path:
            return

        wb = openpyxl.Workbook()

        ws_summary = wb.active
        ws_summary.title = "Summary"
        ws_summary.append(["항목", "값"])
        aggregator = snapshot.get("aggregator", {})
        ws_summary.append(["labor_total", aggregator.get("labor_total", 0)])
        ws_summary.append(["fixed_expense_total", aggregator.get("fixed_expense_total", 0)])
        ws_summary.append(["variable_expense_total", aggregator.get("variable_expense_total", 0)])
        ws_summary.append(["passthrough_expense_total", aggregator.get("passthrough_expense_total", 0)])
        ws_summary.append(["overhead_cost", aggregator.get("overhead_cost", 0)])
        ws_summary.append(["profit", aggregator.get("profit", 0)])
        ws_summary.append(["grand_total", aggregator.get("grand_total", 0)])

        ws_labor = wb.create_sheet("Labor Detail")
        ws_labor.append(
            ["role", "headcount", "base_salary", "bonus", "allowances", "retirement", "labor_subtotal", "role_total"]
        )
        for row in snapshot.get("labor_rows", []):
            ws_labor.append(
                [
                    row["role"],
                    row["headcount"],
                    row["base_salary"],
                    row.get("bonus", 0),
                    row["allowances"],
                    row.get("retirement", 0),
                    row["labor_subtotal"],
                    row["role_total"],
                ]
            )

        ws_expense = wb.create_sheet("Expense Detail")
        ws_expense.append(["category", "name", "driver", "unit_cost", "quantity", "row_total", "type"])
        for row in snapshot.get("expense_rows", []):
            ws_expense.append(
                [
                    row["category"],
                    row["name"],
                    row["driver"],
                    row["unit_cost"],
                    row["quantity"],
                    row["row_total"],
                    row["type"],
                ]
            )

        ws_job = wb.create_sheet("노무비_직종상세")
        ws_job.append(
            [
                "job_code",
                "job_name",
                "headcount",
                "work_days",
                "base_wage",
                "allowance",
                "overtime",
                "total",
            ]
        )
        for row in build_job_breakdown_rows(snapshot):
            ws_job.append(row)

        wb.save(file_path)
        QMessageBox.information(self, "내보내기", "Excel 저장이 완료되었습니다.")

    def _export_full_excel(self):
        """auto_fm_fin.xlsx 양식으로 전체 Excel 내보내기."""
        snapshot = get_result_snapshot(self.last_scenario_id)
        if snapshot is None:
            QMessageBox.information(self, "내보내기", "계산 결과가 없습니다. 먼저 집계를 실행하세요.")
            return

        try:
            from src.domain.export.excel_exporter import ExcelExporter
        except ImportError:
            QMessageBox.warning(self, "내보내기 실패", "export 모듈을 로드할 수 없습니다.")
            return

        template_path = Path(__file__).resolve().parents[1] / "auto_fm_fin.xlsx"
        if not template_path.exists():
            QMessageBox.warning(self, "내보내기 실패", f"템플릿 파일을 찾을 수 없습니다:\n{template_path}")
            return

        input_values = self.input_panel.get_values()
        overhead_rate = input_values.get("overhead_rate", 9.0)
        profit_rate = input_values.get("profit_rate", 10.0)

        safe_name = self._sanitize_filename(self.last_scenario_name)
        default_path = self.scenario_dir / f"{safe_name}_용역원가.xlsx"
        file_path, _ = QFileDialog.getSaveFileName(
            self, "양식 Excel 내보내기", str(default_path), "Excel Files (*.xlsx)",
        )
        if not file_path:
            return

        try:
            exporter = ExcelExporter(
                template_path=template_path,
                snapshot=snapshot,
                overhead_rate=overhead_rate,
                profit_rate=profit_rate,
            )
            exporter.export(Path(file_path))
            QMessageBox.information(self, "내보내기", f"양식 Excel 저장이 완료되었습니다.\n{file_path}")
        except Exception as e:
            logging.getLogger(__name__).error("양식 Excel 내보내기 실패: %s", e, exc_info=True)
            QMessageBox.warning(self, "내보내기 실패", f"Excel 내보내기 중 오류 발생:\n{e}")

    def _export_detail_pdf(self, file_path: str, snapshot: dict) -> None:
        if not file_path:
            return

        printer = QPrinter()
        printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
        printer.setOutputFileName(file_path)

        painter = QPainter(printer)
        y = 40
        line_height = 22

        def draw_line(text: str):
            nonlocal y
            painter.drawText(40, y, text)
            y += line_height

        draw_line(f"시나리오: {self.last_scenario_name}")
        draw_line("노무비 직종 상세")
        draw_line("job_code | job_name | headcount | work_days | base_wage | allowance | overtime | total")

        for row in build_detail_job_rows(snapshot):
            draw_line(" | ".join(str(v) for v in row))

        painter.end()

    # ── 입력 복원 ──

    def _apply_canonical_input(self, canonical: dict) -> None:
        self.job_role_table.table.blockSignals(True)
        try:
            labor_inputs = canonical.get("labor", {}).get("job_roles", {})
            # 저장된 데이터가 있을 때만 적용 (빈 데이터로 덮어쓰지 않음)
            if labor_inputs:
                self.job_role_table.set_job_inputs(labor_inputs)

            # Restore overhead_rate and profit_rate (None 값은 기본값 10으로 처리)
            overhead_rate = canonical.get("overhead_rate")
            if overhead_rate is None:
                overhead_rate = 10
            profit_rate = canonical.get("profit_rate")
            if profit_rate is None:
                profit_rate = 10
            self.input_panel.set_values({
                "scenario_name": self.input_panel.scenario_name.text(),
                "overhead_rate": overhead_rate,
                "profit_rate": profit_rate,
            })
            # 기준연도·노임단가 기준년도·반기 복원 (저장된 값이 있을 때만)
            year_values = {}
            if canonical.get("base_year") is not None:
                year_values["base_year"] = int(canonical["base_year"])
            if canonical.get("wage_year") is not None:
                year_values["wage_year"] = int(canonical["wage_year"])
            if canonical.get("wage_half") is not None:
                year_values["wage_half"] = str(canonical["wage_half"])
            if year_values:
                self.base_year_panel.set_values(year_values)
            # 휴일근무일수 계산 탭 복원
            hw_calc = canonical.get("holiday_work_days_calc")
            if isinstance(hw_calc, dict):
                self.holiday_work_days_panel.set_values({
                    "year": hw_calc.get("year"),
                    "public_holidays_by_month": hw_calc.get("public_holidays_by_month"),
                    "statutory_holidays": hw_calc.get("statutory_holidays", 14),
                    "substitute_holidays": hw_calc.get("substitute_holidays", 0),
                    "center_count": hw_calc.get("center_count", 1),
                    "shift_type": hw_calc.get("shift_type", "미운영"),
                    "crew_size_3shift": hw_calc.get("crew_size_3shift", 1),
                    "headcount_excl_manager": hw_calc.get("headcount_excl_manager", 0),
                    "monthly_work_days": hw_calc.get("monthly_work_days", 20.6),
                    "annual_holiday_work_days": hw_calc.get("annual_holiday_work_days", 0),
                })
                self._sync_holiday_headcount()
            # 불러오기 후 직무 테이블 편집 가능 상태 명시 복원 (시그널 블록 안에서 실행)
            self.job_role_table._force_editable()
        finally:
            self.job_role_table.table.blockSignals(False)

    def _sync_holiday_headcount(self) -> None:
        """직무별 인원에서 관리소장을 제외한 인원을 계산해 휴일근무일수 패널에 반영."""
        job_inputs = self.job_role_table.get_job_inputs()
        role_names = getattr(self, "_role_name_map", {}) or {}
        total = sum(v.get("headcount", 0) for v in job_inputs.values())
        manager_headcount = sum(
            v.get("headcount", 0)
            for job_code, v in job_inputs.items()
            if role_names.get(job_code) == "관리소장" or job_code == "MGR01"
        )
        self.holiday_work_days_panel.set_headcount_excluding_manager(max(0, total - manager_headcount))

    # ── 더티 상태 관리 ──

    def _mark_dirty(self):
        self._set_dirty(True)

    def _on_job_role_changed(self):
        """직무별 인원 입력 변경 시 디바운스 후 실제 계산 수행."""
        self._mark_dirty()
        self._job_role_debounce_timer.start(300)

    def _clear_restoring_snapshot(self) -> None:
        """불러오기 후 일정 시간이 지나면 플래그 해제 (이후 사용자 편집 시 자동계산 정상 동작)."""
        self._restoring_snapshot = False

    def _debounce_expense_detail_refresh(self) -> None:
        """경비입력 수량/단가 셀 변경 시 디바운스 후 경비상세 자동 재계산."""
        self._expense_detail_debounce_timer.stop()
        self._expense_detail_debounce_timer.start(400)

    def _on_expense_quantity_or_price_changed(self) -> None:
        """디바운스 타임아웃: 테이블 → 메모리 반영, 경비상세 재계산, 현재 경비코드 DB 자동 저장."""
        logging.info("[경비입력→경비상세] 수량/단가 변경 감지 → 경비상세 자동 재계산 시작")
        commit_table_edit(self.expense_sub_item_table.table)
        total_headcount = sum(
            v.get("headcount", 0) for v in self.job_role_table.get_job_inputs().values()
        ) or 1
        self.expense_sub_item_table._total_headcount = max(int(total_headcount), 1)
        self.expense_sub_item_table._save_table_to_current_exp()
        self._refresh_expense_detail_from_input()
        self._save_current_expense_only(silent=True)
        logging.info("[경비입력→경비상세] 경비상세 자동 재계산 완료")

    def _on_expense_edit_applied(self) -> None:
        """경비입력 수정 버튼 클릭 시: 상태 메시지 + 경비상세 탭 재계산."""
        self.status_bar.showMessage(
            "선택한 경비코드 수정 내용이 반영되었습니다. '시나리오 저장'을 누르면 저장됩니다.", 4000
        )
        self._refresh_expense_detail_from_input()

    def _refresh_expense_detail_from_input(self) -> None:
        """현재 경비입력(_sub_items_by_exp) 기준으로 경비상세 탭만 재계산해 갱신. DB/직무 변경 없음."""
        commit_table_edit(self.expense_sub_item_table.table)
        values = self.input_panel.get_values()
        scenario_name = (values.get("scenario_name") or "").strip() or "default"
        scenario_id = self._sanitize_filename(scenario_name)
        if not scenario_id:
            scenario_id = "default"
        est = self.expense_sub_item_table
        if not est._sub_items_by_exp:
            return
        try:
            conn = get_connection()
            try:
                job_inputs = self.job_role_table.get_job_inputs()
                year_values = self.base_year_panel.get_values()
                wage_year = year_values.get("wage_year")
                labor_rows = get_labor_rows_from_ui(
                    job_inputs, scenario_id, conn, wage_year=wage_year
                )
                labor_total = sum(
                    int(r.get("role_total", 0)) for r in labor_rows if isinstance(r, dict)
                )
                expense_rows = get_expense_rows_for_display(
                    scenario_id, conn, est._sub_items_by_exp, labor_total
                )
                total_headcount = sum(
                    v.get("headcount", 0) for v in job_inputs.values()
                ) or 1
                self.last_expense_rows = expense_rows
                self.expense_detail.update_rows(expense_rows, total_headcount=total_headcount)
                self.expense_detail.update()
            finally:
                conn.close()
        except Exception as exc:
            logging.exception("경비상세 재계산 중 오류: %s", exc)

    def _do_job_role_changed(self):
        """직무별 인원 입력이 변경되면 보험료를 자동 계산하여 경비입력에 반영. DB 반영 없이 현재 UI 기준으로만 계산."""
        if getattr(self, "_restoring_snapshot", False):
            # 플래그를 여기서 해제하지 않음 → 400ms 타이머가 해제 (보호 기간 내 중복 호출 방지)
            return
        values = self.input_panel.get_values()
        scenario_name = values.get("scenario_name") or "default"

        scenario_id = self._sanitize_filename(scenario_name)

        if not scenario_id:
            scenario_id = "default"

        commit_table_edit(self.job_role_table.table)
        job_inputs = self.job_role_table.get_job_inputs()

        # 기준년도 패널에서 노임단가 기준년도 읽기 (시나리오별 올바른 임금 기준 적용)
        year_values = self.base_year_panel.get_values()
        wage_year = year_values.get("wage_year")

        try:
            conn = get_connection()
            try:
                # 노무비 상세 계산 및 표시
                labor_rows = get_labor_rows_from_ui(job_inputs, scenario_id, conn, wage_year=wage_year)
                # ISSUE-008 수정: 집계 결과가 있을 때는 빈 데이터로 덮어쓰지 않음
                if labor_rows or not self.last_labor_rows:
                    self.last_labor_rows = labor_rows
                    display_rows = labor_rows
                else:
                    # 빈 데이터면 기존 집계 결과 유지
                    display_rows = self.last_labor_rows
                self.labor_detail.update_rows(display_rows)
                self.labor_detail.update()
                n_labor = len(display_rows)
                if n_labor != self._last_labor_count:
                    logging.info("노무비 상세 자동계산 완료: %s개 직무 반영", n_labor)
                    self._last_labor_count = n_labor

                labor_total = sum(
                    int(r.get("role_total", 0)) for r in labor_rows if isinstance(r, dict)
                )
                aggregator = Aggregator(
                    labor_total, 0, 0, 0, 0, 0
                )
                self.last_aggregator = aggregator
                self.summary_panel.update_summary(aggregator, pdf_grand_total=0)
                self.summary_panel.update()
                self.donut_chart.update_aggregator(aggregator)
                self._refresh_button_state()

                # 직무별 인원 → 노무비 보험료 7종 → 경비입력 경비코드(보험 7종) 자동 매핑
                labor_insurance = get_insurance_by_exp_code_from_ui(job_inputs, scenario_id, conn, wage_year=wage_year) or {}
                n_ins = len(labor_insurance)
                if n_ins != self._last_insurance_count:
                    logging.info("보험료 자동계산 완료: %s개 항목 반영", n_ins)
                    self._last_insurance_count = n_ins

                # 경비 항목 정보 가져오기 (새 시나리오면 default에서 경비코드 복사)
                repo = MasterDataRepo(conn)
                repo.ensure_expense_masterdata_for_scenario(scenario_id)
                items = repo.get_expense_items(scenario_id)
                expense_items_for_sub = [
                    {
                        "exp_code": i.exp_code,
                        "exp_name": i.exp_name,
                        "group_code": i.group_code,
                        "sort_order": i.sort_order,
                    }
                    for i in items
                ]

                # 총 인원 수 계산
                total_headcount = sum(
                    v.get("headcount", 0) for v in job_inputs.values()
                )

                # 사용자 편집 중인 비보험 항목 보존: 보험 7종만 DB 재계산으로 갱신
                est = self.expense_sub_item_table

                # 현재 테이블 편집 내용을 먼저 in-memory에 저장
                commit_table_edit(est.table)
                est._save_table_to_current_exp()
                user_edits = dict(est._sub_items_by_exp)

                # DB에서 세부 항목 재구성 (보험 7종 포함)
                sub_items_by_exp = build_sub_items_by_exp(
                    repo, scenario_id,
                    labor_insurance_by_exp_code=labor_insurance,
                    expense_items=expense_items_for_sub,
                )

                # 사용자 편집 병합: 비보험 코드는 사용자 편집값 유지
                for code, items in user_edits.items():
                    if code not in EXP_CODES_FROM_LABOR and items:
                        sub_items_by_exp[code] = items

                # 경비입력 탭이 보일 때만 테이블 UI 갱신, 아닐 때도 in-memory 데이터는 업데이트
                if self.tabs.currentIndex() == self.tabs.indexOf(self.expense_sub_item_table):
                    est.table.blockSignals(True)
                    try:
                        est.load_sub_items(
                            sub_items_by_exp, expense_items_for_sub, total_headcount=total_headcount
                        )
                    finally:
                        est.table.blockSignals(False)
                else:
                    # 다른 탭에 있을 때: in-memory 데이터만 갱신 (보험 7종 반영 + 사용자 편집 보존)
                    est._sub_items_by_exp = {
                        exp_code: [_sub_item_to_dict(si) for si in items]
                        for exp_code, items in sub_items_by_exp.items()
                    }

                # 경비 상세에도 보험 7종 포함해 반영
                expense_rows = get_expense_rows_for_display(
                    scenario_id, conn, sub_items_by_exp, labor_total
                )
                self.last_expense_rows = expense_rows
                self.expense_detail.update_rows(expense_rows, total_headcount=total_headcount)
                self.expense_detail.update()

                self.status_bar.showMessage(f"✓ 노무비 자동계산 완료 (인원: {total_headcount}명)", 3000)

                # 휴일근무일수 탭: 관리소장 제외 인원 자동 반영
                self._sync_holiday_headcount()

            finally:
                conn.close()
        except Exception as exc:
            error_msg = f"노무비 자동계산 중 오류: {str(exc)}"
            logging.exception(error_msg)
            self.status_bar.showMessage(f"⚠ {error_msg}", 5000)
            QMessageBox.warning(
                self,
                "자동계산 오류",
                f"{error_msg}\n\n자세한 내용은 로그를 확인하세요."
            )

    def _set_dirty(self, value: bool):
        self._dirty = value
        # 타이틀에 더티 마커 표시
        if value:
            self.setWindowTitle("* 자동집하시설 원가산정 프로그램")
            self.status_bar.showMessage("변경사항 있음")
        else:
            self.setWindowTitle("자동집하시설 원가산정 프로그램")
            self.status_bar.showMessage("저장됨")
        self._refresh_button_state()

    def _refresh_button_state(self):
        if _APP_CONTROLLERS_AVAILABLE and self._ctx is not None:
            has_snapshot = self._ctx.has_snapshot()
            not_editing = not (
                self.job_role_table.is_editing()
                or self.expense_sub_item_table.table.state() == QAbstractItemView.State.EditingState
            )
            # 시나리오 저장: 입력만 있어도 저장 가능(집계 없이). 편집 중이 아닐 때만 버튼 활성화
            self.save_button.setEnabled(not_editing)
            self.calculate_button.setEnabled(not_editing)
            self.export_pdf_button.setEnabled(has_snapshot)
            self.export_excel_button.setEnabled(has_snapshot)
            self.export_full_excel_button.setEnabled(has_snapshot)
        else:
            state = compute_action_state(self._dirty, self.last_aggregator is not None)
            self.save_button.setEnabled(state["save_enabled"])
            self.calculate_button.setEnabled(state["calculate_enabled"])
            self.export_pdf_button.setEnabled(state["export_enabled"])
            self.export_excel_button.setEnabled(state["export_enabled"])
            self.export_full_excel_button.setEnabled(state["export_enabled"])

    def _show_validation_errors(self, errors):
        lines = [f"{e.field_path}: {e.message}" for e in errors]
        QMessageBox.warning(self, "검증 오류", "\n".join(lines))
