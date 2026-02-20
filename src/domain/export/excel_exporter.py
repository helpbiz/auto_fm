"""auto_fm_fin.xlsx 템플릿 기반 Excel 내보내기."""

import logging
from pathlib import Path

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .cell_maps import (
    GAPJI_CELLS,
    SERVICE_COST_SUMMARY_CELLS,
    YEARLY_COMMON_ROWS,
    YEARLY_BOTTOM_OFFSETS,
    YEARLY_VAR_KEYS_TO_EXP_CODE,
    YEARLY_PASS_KEYS_TO_EXP_CODE,
    LABOR_SUMMARY_COLS,
    LABOR_DATA_START_ROW,
    LABOR_TOTAL_ROW,
    EXPENSE_SUMMARY_ROW_MAP,
    EXPENSE_SUMMARY_COL,
    EXPENSE_INS_SUBTOTAL_ROW,
    EXPENSE_FIXED_TOTAL_ROW,
    EXPENSE_VAR_TOTAL_ROW,
    EXPENSE_PASS_TOTAL_ROW,
    EXPENSE_GRAND_TOTAL_ROW,
    OVERHEAD_CELLS,
    PROFIT_CELLS,
)

log = logging.getLogger(__name__)


class ExcelExporter:
    """시나리오 집계 결과를 auto_fm_fin.xlsx 양식으로 내보내기."""

    def __init__(
        self,
        template_path: Path,
        snapshot: dict,
        overhead_rate: float = 9.0,
        profit_rate: float = 10.0,
    ):
        self.wb: Workbook = openpyxl.load_workbook(str(template_path))
        self.snapshot = snapshot
        self.overhead_rate = overhead_rate
        self.profit_rate = profit_rate

        agg = snapshot.get("aggregator", {})
        self._labor_total: int = int(agg.get("labor_total", 0))
        self._fixed_total: int = int(agg.get("fixed_expense_total", 0))
        self._variable_total: int = int(agg.get("variable_expense_total", 0))
        self._passthrough_total: int = int(agg.get("passthrough_expense_total", 0))
        self._overhead_cost: int = int(agg.get("overhead_cost", 0))
        self._profit: int = int(agg.get("profit", 0))
        self._grand_total: int = int(agg.get("grand_total", 0))

        self._labor_rows: list[dict] = snapshot.get("labor_rows", [])
        self._expense_rows: list[dict] = snapshot.get("expense_rows", [])
        self._insurance: dict[str, int] = snapshot.get("insurance_by_exp_code", {})

        # exp_code → monthly amount 캐시
        self._expense_map: dict[str, int] = {}
        for row in self._expense_rows:
            code = row.get("exp_code", "")
            try:
                amt = int(float(row.get("row_total", 0)))
            except (ValueError, TypeError):
                amt = 0
            self._expense_map[code] = amt

    # ------------------------------------------------------------------
    # public
    # ------------------------------------------------------------------

    def export(self, output_path: Path) -> Path:
        """Phase 1 시트에 데이터 주입 후 저장."""
        self._write_gapji()
        self._write_service_cost_summary()
        for sheet_name in YEARLY_BOTTOM_OFFSETS:
            self._write_yearly_cost_sheet(sheet_name)
        self._write_labor_summary()
        self._write_expense_summary()
        self._write_overhead()
        self._write_profit()
        self.wb.save(str(output_path))
        log.info("Excel exported to %s", output_path)
        return output_path

    # ------------------------------------------------------------------
    # helpers
    # ------------------------------------------------------------------

    def _get_ws(self, name: str) -> Worksheet:
        """시트명으로 worksheet 반환 (공백 트림 대응)."""
        for sn in self.wb.sheetnames:
            if sn.strip() == name.strip():
                return self.wb[sn]
        raise KeyError(f"Sheet not found: {name}")

    def _set(self, ws: Worksheet, coord: str, value) -> None:
        """기존 서식을 보존하면서 값만 설정."""
        ws[coord].value = value

    def _expense_monthly(self, exp_code: str) -> int:
        """단일 exp_code의 월 합계."""
        return self._expense_map.get(exp_code, 0)

    def _expense_monthly_sum(self, exp_codes: list[str]) -> int:
        """여러 exp_code의 월 합계."""
        return sum(self._expense_map.get(c, 0) for c in exp_codes)

    def _insurance_monthly(self, exp_code: str) -> int:
        return int(self._insurance.get(exp_code, 0))

    def _truncate_1000(self, value: int) -> int:
        """천원 단위 절사."""
        return (value // 1000) * 1000

    # ------------------------------------------------------------------
    # 갑지
    # ------------------------------------------------------------------

    def _write_gapji(self) -> None:
        try:
            ws = self._get_ws("갑지")
        except KeyError:
            log.warning("갑지 시트를 찾을 수 없습니다")
            return

        annual = self._grand_total * 12
        annual_truncated = self._truncate_1000(annual)
        total_3year = annual_truncated * 3

        self._set(ws, GAPJI_CELLS["total_3year"], total_3year)
        self._set(ws, GAPJI_CELLS["year1"], annual_truncated)
        self._set(ws, GAPJI_CELLS["year2"], annual_truncated)
        self._set(ws, GAPJI_CELLS["year3"], annual_truncated)

    # ------------------------------------------------------------------
    # 용역원가집계
    # ------------------------------------------------------------------

    def _write_service_cost_summary(self) -> None:
        try:
            ws = self._get_ws("용역원가집계")
        except KeyError:
            log.warning("용역원가집계 시트를 찾을 수 없습니다")
            return

        c = SERVICE_COST_SUMMARY_CELLS
        monthly = self._grand_total
        annual = monthly * 12

        # 1~3차 동일 복사
        for prefix in ("y1", "y2", "y3"):
            self._set(ws, c[f"{prefix}_monthly"], monthly)
            self._set(ws, c[f"{prefix}_months"], 12)
            self._set(ws, c[f"{prefix}_annual"], annual)

        self._set(ws, c["grand_total"], annual * 3)

    # ------------------------------------------------------------------
    # 용역원가(XX년) — 연도별 원가계산서
    # ------------------------------------------------------------------

    def _write_yearly_cost_sheet(self, sheet_name: str) -> None:
        try:
            ws = self._get_ws(sheet_name)
        except KeyError:
            log.warning("%s 시트를 찾을 수 없습니다", sheet_name)
            return

        offsets = YEARLY_BOTTOM_OFFSETS.get(sheet_name)
        if offsets is None:
            log.warning("행 오프셋 정의 없음: %s", sheet_name)
            return

        # -- 상단: 인건비 --
        base_salary_sum = sum(r.get("base_salary", 0) for r in self._labor_rows)
        allowances_sum = sum(r.get("allowances", 0) for r in self._labor_rows)
        bonus_sum = sum(r.get("bonus", 0) for r in self._labor_rows)
        retirement_sum = sum(r.get("retirement", 0) for r in self._labor_rows)

        def _set_gh(row: int, monthly: int) -> None:
            self._set(ws, f"G{row}", monthly)
            self._set(ws, f"H{row}", monthly * 12)

        _set_gh(YEARLY_COMMON_ROWS["base_salary"], base_salary_sum)
        _set_gh(YEARLY_COMMON_ROWS["allowances_subtotal"], allowances_sum)
        _set_gh(YEARLY_COMMON_ROWS["bonus"], bonus_sum)
        _set_gh(YEARLY_COMMON_ROWS["retirement"], retirement_sum)
        _set_gh(YEARLY_COMMON_ROWS["labor_total"], self._labor_total)

        # -- 보험료 7종 --
        ins_map = {
            "ins_indust": "FIX_INS_INDUST",
            "ins_pension": "FIX_INS_PENSION",
            "ins_employ": "FIX_INS_EMPLOY",
            "ins_health": "FIX_INS_HEALTH",
            "ins_longterm": "FIX_INS_LONGTERM",
            "ins_wage": "FIX_INS_WAGE",
            "ins_asbestos": "FIX_INS_ASBESTOS",
        }
        for key, exp_code in ins_map.items():
            row = YEARLY_COMMON_ROWS[key]
            val = self._insurance_monthly(exp_code)
            _set_gh(row, val)

        # -- 복리후생비 --
        welfare_codes = ["FIX_WEL_CLOTH", "FIX_WEL_MEAL", "FIX_WEL_CHECKUP", "FIX_WEL_MEDICINE"]
        welfare_sum = self._expense_monthly_sum(welfare_codes)
        _set_gh(YEARLY_COMMON_ROWS["welfare"], welfare_sum)

        # -- 기타 고정경비 --
        fixed_exp_map = {
            "safety": "FIX_SAFETY",
            "training": "FIX_TRAINING",
            "supplies": "FIX_SUPPLIES",
            "travel": "FIX_TRAVEL",
            "telecom": "FIX_TELECOM",
        }
        for key, exp_code in fixed_exp_map.items():
            row = YEARLY_COMMON_ROWS[key]
            _set_gh(row, self._expense_monthly(exp_code))

        # -- 하단: 고정경비소계~총계 --
        _set_gh(offsets["fixed_subtotal"], self._fixed_total)

        net_service = self._labor_total + self._fixed_total
        _set_gh(offsets["net_service"], net_service)
        _set_gh(offsets["overhead"], self._overhead_cost)
        _set_gh(offsets["profit"], self._profit)

        service_cost = net_service + self._overhead_cost + self._profit
        _set_gh(offsets["service_cost"], service_cost)

        contingency = int(service_cost * 0.10)
        _set_gh(offsets["contingency"], contingency)

        fixed_grand = service_cost + contingency
        _set_gh(offsets["fixed_grand"], fixed_grand)

        # -- 변동경비 --
        for var_key, exp_code in YEARLY_VAR_KEYS_TO_EXP_CODE:
            row = offsets[var_key]
            _set_gh(row, self._expense_monthly(exp_code))

        _set_gh(offsets["var_subtotal"], self._variable_total)

        var_service = self._variable_total
        _set_gh(offsets["var_service"], var_service)

        var_contingency = int(var_service * 0.10)
        _set_gh(offsets["var_contingency"], var_contingency)

        var_grand = var_service + var_contingency
        _set_gh(offsets["var_grand"], var_grand)

        # -- 대행비 --
        for pass_key, exp_code in YEARLY_PASS_KEYS_TO_EXP_CODE:
            row = offsets[pass_key]
            _set_gh(row, self._expense_monthly(exp_code))

        _set_gh(offsets["pass_subtotal"], self._passthrough_total)

        all_grand = fixed_grand + var_grand
        _set_gh(offsets["all_grand"], all_grand)

        total = self._grand_total
        _set_gh(offsets["total"], total)

    # ------------------------------------------------------------------
    # 인건비집계
    # ------------------------------------------------------------------

    def _write_labor_summary(self) -> None:
        try:
            ws = self._get_ws("인건비집계")
        except KeyError:
            log.warning("인건비집계 시트를 찾을 수 없습니다")
            return

        cols = LABOR_SUMMARY_COLS
        active_rows = [r for r in self._labor_rows if r.get("headcount", 0) > 0]

        # 합계 누적
        sum_base = 0
        sum_allow = 0
        sum_bonus = 0
        sum_retire = 0
        sum_total = 0
        sum_headcount = 0

        for i, lr in enumerate(active_rows):
            row_num = LABOR_DATA_START_ROW + i
            if row_num >= LABOR_TOTAL_ROW:
                break

            hc = lr.get("headcount", 0)
            base = lr.get("base_salary", 0)
            allow = lr.get("allowances", 0)
            bonus = lr.get("bonus", 0)
            retire = lr.get("retirement", 0)
            subtotal = lr.get("labor_subtotal", 0)

            base_pp = int(base / hc) if hc else 0
            allow_pp = int(allow / hc) if hc else 0
            bonus_pp = int(bonus / hc) if hc else 0
            retire_pp = int(retire / hc) if hc else 0
            total_pp = int(subtotal / hc) if hc else 0

            self._set(ws, f"{cols['base_per_person']}{row_num}", base_pp)
            self._set(ws, f"{cols['base_headcount']}{row_num}", hc)
            self._set(ws, f"{cols['base_amount']}{row_num}", base)

            self._set(ws, f"{cols['allow_per_person']}{row_num}", allow_pp)
            self._set(ws, f"{cols['allow_headcount']}{row_num}", hc)
            self._set(ws, f"{cols['allow_amount']}{row_num}", allow)

            self._set(ws, f"{cols['bonus_per_person']}{row_num}", bonus_pp)
            self._set(ws, f"{cols['bonus_headcount']}{row_num}", hc)
            self._set(ws, f"{cols['bonus_amount']}{row_num}", bonus)

            self._set(ws, f"{cols['retire_per_person']}{row_num}", retire_pp)
            self._set(ws, f"{cols['retire_headcount']}{row_num}", hc)
            self._set(ws, f"{cols['retire_amount']}{row_num}", retire)

            self._set(ws, f"{cols['total_per_person']}{row_num}", total_pp)
            self._set(ws, f"{cols['total_headcount']}{row_num}", hc)
            self._set(ws, f"{cols['total_amount']}{row_num}", subtotal)

            sum_base += base
            sum_allow += allow
            sum_bonus += bonus
            sum_retire += retire
            sum_total += subtotal
            sum_headcount += hc

        # 합계 행
        tr = LABOR_TOTAL_ROW
        self._set(ws, f"{cols['base_headcount']}{tr}", sum_headcount)
        self._set(ws, f"{cols['base_amount']}{tr}", sum_base)
        self._set(ws, f"{cols['allow_headcount']}{tr}", sum_headcount)
        self._set(ws, f"{cols['allow_amount']}{tr}", sum_allow)
        self._set(ws, f"{cols['bonus_headcount']}{tr}", sum_headcount)
        self._set(ws, f"{cols['bonus_amount']}{tr}", sum_bonus)
        self._set(ws, f"{cols['retire_headcount']}{tr}", sum_headcount)
        self._set(ws, f"{cols['retire_amount']}{tr}", sum_retire)
        self._set(ws, f"{cols['total_headcount']}{tr}", sum_headcount)
        self._set(ws, f"{cols['total_amount']}{tr}", sum_total)

    # ------------------------------------------------------------------
    # 경비집계
    # ------------------------------------------------------------------

    def _write_expense_summary(self) -> None:
        try:
            ws = self._get_ws("경비집계")
        except KeyError:
            log.warning("경비집계 시트를 찾을 수 없습니다")
            return

        col = EXPENSE_SUMMARY_COL

        insurance_subtotal = 0
        for row_num, exp_code_or_list in EXPENSE_SUMMARY_ROW_MAP.items():
            if isinstance(exp_code_or_list, list):
                val = self._expense_monthly_sum(exp_code_or_list)
            elif exp_code_or_list.startswith("FIX_INS_"):
                val = self._insurance_monthly(exp_code_or_list)
            else:
                val = self._expense_monthly(exp_code_or_list)

            self._set(ws, f"{col}{row_num}", val)

            # 보험료 소계 누적 (행 7~13)
            if 7 <= row_num <= 13:
                insurance_subtotal += val

        # 보험료 소계 (행 14)
        self._set(ws, f"{col}{EXPENSE_INS_SUBTOTAL_ROW}", insurance_subtotal)

        # 고정경비계 (행 22)
        self._set(ws, f"{col}{EXPENSE_FIXED_TOTAL_ROW}", self._fixed_total)

        # 변동경비계 (행 30)
        self._set(ws, f"{col}{EXPENSE_VAR_TOTAL_ROW}", self._variable_total)

        # 대행비계 (행 34)
        self._set(ws, f"{col}{EXPENSE_PASS_TOTAL_ROW}", self._passthrough_total)

        # 합계 (행 35)
        total = self._fixed_total + self._variable_total + self._passthrough_total
        self._set(ws, f"{col}{EXPENSE_GRAND_TOTAL_ROW}", total)

    # ------------------------------------------------------------------
    # 일반관리비
    # ------------------------------------------------------------------

    def _write_overhead(self) -> None:
        try:
            ws = self._get_ws("일반관리비")
        except KeyError:
            log.warning("일반관리비 시트를 찾을 수 없습니다")
            return

        c = OVERHEAD_CELLS
        self._set(ws, c["labor"], self._labor_total)
        self._set(ws, c["fixed"], self._fixed_total)
        self._set(ws, c["sum"], self._labor_total + self._fixed_total)
        self._set(ws, c["rate"], self.overhead_rate)
        self._set(ws, c["amount"], self._overhead_cost)
        self._set(ws, c["total"], self._overhead_cost)

    # ------------------------------------------------------------------
    # 이윤
    # ------------------------------------------------------------------

    def _write_profit(self) -> None:
        try:
            ws = self._get_ws("이윤")
        except KeyError:
            log.warning("이윤 시트를 찾을 수 없습니다")
            return

        c = PROFIT_CELLS
        self._set(ws, c["labor"], self._labor_total)
        self._set(ws, c["fixed"], self._fixed_total)
        self._set(ws, c["overhead"], self._overhead_cost)
        profit_base = self._labor_total + self._fixed_total + self._overhead_cost
        self._set(ws, c["sum"], profit_base)
        self._set(ws, c["rate"], self.profit_rate)
        self._set(ws, c["amount"], self._profit)
        self._set(ws, c["total"], self._profit)
